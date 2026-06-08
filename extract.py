#!/usr/bin/env python3
"""
Project Analyzer - Generates PROJECT_TREE.txt and PROJECT_ANALYSIS.xlsx
Run from project root directory.
"""

import os
import sys
import ast
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             GradientFill)
from openpyxl.utils import get_column_letter

# ── Exclusions ────────────────────────────────────────────────────────────────
EXCLUDE_FOLDERS = {
    '__pycache__', '.git', '.venv', 'venv', 'env', 'node_modules',
    'build', 'dist', '.idea', '.vscode', '.pytest_cache', '.mypy_cache',
    '.coverage', 'htmlcov', '.tox', '.eggs',
}
EXCLUDE_FILES = {
    '*.pyc', '*.pyo', '*.pyd', '*.so', '*.dll', '*.exe',
    '*.log', '*.tmp', '*.temp', '*.swp', '*.swo',
    '*.bak', '*.orig', '*.class', '*.o', '*.obj',
    '.DS_Store', 'Thumbs.db', 'desktop.ini',
    'PROJECT_TREE.txt', 'PROJECT_ANALYSIS.xlsx',
    '*.egg', '*.egg-info', '*.cache', '*.lock',
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def should_exclude(name, is_dir=False):
    if is_dir:
        return name in EXCLUDE_FOLDERS
    for pat in EXCLUDE_FILES:
        if pat.startswith('*') and name.endswith(pat[1:]):
            return True
        elif pat.endswith('*') and name.startswith(pat[:-1]):
            return True
        elif pat == name:
            return True
    return False


def get_file_type(name):
    ext = Path(name).suffix.lower()
    mapping = {
        '.py': 'Python', '.js': 'JavaScript', '.ts': 'TypeScript',
        '.jsx': 'React JSX', '.tsx': 'React TSX', '.html': 'HTML',
        '.css': 'CSS', '.scss': 'SCSS', '.json': 'JSON', '.yaml': 'YAML',
        '.yml': 'YAML', '.toml': 'TOML', '.txt': 'Text', '.md': 'Markdown',
        '.sql': 'SQL', '.sh': 'Shell', '.bat': 'Batch', '.env': 'Env',
        '.xml': 'XML', '.csv': 'CSV', '.ini': 'Config', '.cfg': 'Config',
    }
    return mapping.get(ext, ext[1:].upper() if ext else 'Unknown')


# ── Tree Generator ─────────────────────────────────────────────────────────────

def generate_tree(start_path, prefix=""):
    lines = []
    try:
        items = sorted(os.listdir(start_path))
    except PermissionError:
        return lines

    dirs, files = [], []
    for item in items:
        item_path = os.path.join(start_path, item)
        if should_exclude(item, os.path.isdir(item_path)):
            continue
        (dirs if os.path.isdir(item_path) else files).append(item)

    for f in files:
        lines.append(f"{prefix}|   {f}")
    for i, d in enumerate(dirs):
        last = i == len(dirs) - 1
        # Fixed: Use chr(92) for backslash or double backslash
        if last:
            lines.append(f"{prefix}\\---{d}")
        else:
            lines.append(f"{prefix}+---{d}")
        lines.extend(generate_tree(os.path.join(start_path, d),
                                   prefix + ("    " if last else "|   ")))
    return lines


def write_tree_txt(project_root, out_file="PROJECT_TREE.txt"):
    tree = generate_tree(project_root)
    lines = [f"PROJECT ROOT : {project_root}", "|"] + tree + ["", "===== END OF TREE ====="]
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"[OK] PROJECT_TREE.txt written")
    return lines


# ── File Scanner ───────────────────────────────────────────────────────────────

def collect_files(root):
    """Walk project and collect all non-excluded files."""
    records = []
    sno = 1
    for dirpath, dirnames, filenames in os.walk(root):
        # Prune excluded dirs
        dirnames[:] = [d for d in sorted(dirnames) if not should_exclude(d, True)]
        for fname in sorted(filenames):
            if should_exclude(fname, False):
                continue
            full = os.path.join(dirpath, fname)
            rel = os.path.relpath(full, root)
            folder = os.path.relpath(dirpath, root)
            if folder == '.':
                folder = '(root)'
            records.append({
                'sno': sno,
                'folder': folder,
                'file': fname,
                'type': get_file_type(fname),
                'path': rel.replace('\\', '/'),
            })
            sno += 1
    return records


# ── Python AST Parser ─────────────────────────────────────────────────────────

def parse_python_file(filepath):
    """Return (classes, methods, functions) from a Python file."""
    classes, methods, functions = [], [], []

    class Visitor(ast.NodeVisitor):
        def __init__(self):
            self.current_class = None

        def visit_ClassDef(self, node):
            classes.append(node.name)
            prev = self.current_class
            self.current_class = node.name
            self.generic_visit(node)
            self.current_class = prev

        def visit_FunctionDef(self, node):
            if self.current_class:
                methods.append((self.current_class, node.name))
            else:
                functions.append(node.name)
            # Visit nested classes only
            for child in ast.iter_child_nodes(node):
                if isinstance(child, ast.ClassDef):
                    self.visit(child)

        visit_AsyncFunctionDef = visit_FunctionDef

    try:
        src = Path(filepath).read_text(encoding='utf-8', errors='ignore')
        tree = ast.parse(src, filename=filepath)
        Visitor().visit(tree)
    except Exception:
        pass

    return classes, methods, functions


def collect_python_data(file_records, root):
    """For every Python file, parse AST and return class/method/function rows."""
    class_rows, method_rows, func_rows = [], [], []
    c_sno = m_sno = f_sno = 1

    for rec in file_records:
        if rec['type'] != 'Python':
            continue
        abs_path = os.path.join(root, rec['path'].replace('/', os.sep))
        classes, methods, functions = parse_python_file(abs_path)

        fname = rec['file']

        for cls in classes:
            class_rows.append({'sno': c_sno, 'file': fname, 'class': cls})
            c_sno += 1

        for cls, meth in methods:
            method_rows.append({'sno': m_sno, 'file': fname, 'class': cls, 'method': meth})
            m_sno += 1

        for fn in functions:
            func_rows.append({'sno': f_sno, 'file': fname, 'function': fn})
            f_sno += 1

    return class_rows, method_rows, func_rows


# ── Excel Builder ─────────────────────────────────────────────────────────────

# Color palette
C_HEADER_BG = "1F3864"  # dark navy
C_HEADER_FG = "FFFFFF"
C_ALT1 = "EBF0FA"  # light blue-grey
C_ALT2 = "FFFFFF"
C_ACCENT = "2E75B6"  # mid blue
C_GREEN_BG = "E2EFDA"
C_PURPLE_BG = "F0E6FA"
C_ORANGE_BG = "FFF2CC"
C_TEAL_BG = "E2F0F7"

SHEET_HEADER_BG = {
    "Project Files": "2E75B6",
    "Python Classes": "375623",
    "Python Methods": "7030A0",
    "Functions": "BF8F00",
}

SHEET_ROW_ALT = {
    "Project Files": ("EBF0FA", "FFFFFF"),
    "Python Classes": ("E2EFDA", "FFFFFF"),
    "Python Methods": ("F0E6FA", "FFFFFF"),
    "Functions": ("FFF2CC", "FFFFFF"),
}


def hdr_style(bg, fg="FFFFFF"):
    return {
        'font': Font(bold=True, color=fg, name='Arial', size=10),
        'fill': PatternFill("solid", start_color=bg),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            bottom=Side(style='medium', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
        )
    }


def data_style(bg, fg="000000", bold=False):
    return {
        'font': Font(color=fg, name='Arial', size=9, bold=bold),
        'fill': PatternFill("solid", start_color=bg),
        'alignment': Alignment(vertical='center', wrap_text=False),
        'border': Border(
            bottom=Side(style='hair', color='D0D0D0'),
            right=Side(style='hair', color='D0D0D0'),
        )
    }


def apply_style(cell, style):
    for attr, val in style.items():
        setattr(cell, attr, val)


def sno_style(bg):
    return {
        'font': Font(color="666666", name='Arial', size=9),
        'fill': PatternFill("solid", start_color=bg),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            bottom=Side(style='hair', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
        )
    }


def build_sheet(ws, sheet_name, headers, col_widths, rows_data):
    bg_hdr = SHEET_HEADER_BG[sheet_name]
    alt1, alt2 = SHEET_ROW_ALT[sheet_name]

    # Title row
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"PROJECT ANALYSIS - {sheet_name.upper()}")
    title_cell.font = Font(bold=True, color="FFFFFF", name='Arial', size=13)
    title_cell.fill = PatternFill("solid", start_color=bg_hdr)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Header row
    ws.row_dimensions[2].height = 22
    hs = hdr_style(bg_hdr)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        apply_style(cell, hs)

    # Data rows
    for r, row in enumerate(rows_data, 3):
        ws.row_dimensions[r].height = 18
        bg = alt1 if r % 2 == 1 else alt2
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == 1:  # S.No
                apply_style(cell, sno_style(bg))
            else:
                ds = data_style(bg)
                apply_style(cell, ds)

    # Column widths
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze panes
    ws.freeze_panes = ws['A3']

    # Auto-filter
    if rows_data:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(rows_data)}"

    # Row count note
    note_row = 3 + len(rows_data)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    note_cell = ws.cell(row=note_row, column=1,
                        value=f"Total Records: {len(rows_data)}")
    note_cell.font = Font(italic=True, color="666666", name='Arial', size=8)
    note_cell.fill = PatternFill("solid", start_color="F5F5F5")
    note_cell.alignment = Alignment(horizontal='right')


def build_excel(file_records, class_rows, method_rows, func_rows, out_file):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: Project Files ────────────────────────────────────────────────
    ws1 = wb.create_sheet("Project Files")
    headers1 = ["S.No", "Folder", "File Name", "File Type", "Full Path"]
    widths1 = [6, 35, 30, 12, 55]
    rows1 = [(r['sno'], r['folder'], r['file'], r['type'], r['path']) for r in file_records]
    build_sheet(ws1, "Project Files", headers1, widths1, rows1)

    # Colorize file type column (col 4)
    TYPE_COLORS = {
        'Python': ('E8F5E9', '1B5E20'),
        'JavaScript': ('FFF8E1', '7B5E00'),
        'TypeScript': ('E3F2FD', '0D47A1'),
        'HTML': ('FBE9E7', 'B71C1C'),
        'CSS': ('F3E5F5', '4A148C'),
        'JSON': ('E0F2F1', '004D40'),
        'YAML': ('FFF3E0', 'E65100'),
        'Markdown': ('E8EAF6', '1A237E'),
        'Text': ('FAFAFA', '424242'),
        'Shell': ('F9FBE7', '33691E'),
        'SQL': ('EDE7F6', '311B92'),
    }
    for row_idx, rec in enumerate(file_records, 3):
        t = rec['type']
        if t in TYPE_COLORS:
            bg, fg = TYPE_COLORS[t]
            cell = ws1.cell(row=row_idx, column=4)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.font = Font(color=fg, name='Arial', size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Sheet 2: Python Classes ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Python Classes")
    headers2 = ["S.No", "File Name", "Class Name"]
    widths2 = [6, 35, 40]
    rows2 = [(r['sno'], r['file'], r['class']) for r in class_rows]
    build_sheet(ws2, "Python Classes", headers2, widths2, rows2)

    # ── Sheet 3: Python Methods ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Python Methods")
    headers3 = ["S.No", "File Name", "Class Name", "Method Name"]
    widths3 = [6, 35, 35, 35]
    rows3 = [(r['sno'], r['file'], r['class'], r['method']) for r in method_rows]
    build_sheet(ws3, "Python Methods", headers3, widths3, rows3)

    # Extra: alternate row color by class group for visual clarity
    if method_rows:
        prev_class = None
        toggle = False
        ALT_CLASS_COLORS = [("F0E6FA", "E6D3F5"), ("FAFAFA", "F0F0F0")]
        for row_idx, mr in enumerate(method_rows, 3):
            if mr['class'] != prev_class:
                toggle = not toggle
                prev_class = mr['class']
            bg_pair = ALT_CLASS_COLORS[0] if toggle else ALT_CLASS_COLORS[1]
            bg = bg_pair[0] if row_idx % 2 == 1 else bg_pair[1]
            for col in range(1, 5):
                cell = ws3.cell(row=row_idx, column=col)
                cell.fill = PatternFill("solid", start_color=bg)

    # ── Sheet 4: Functions ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Functions")
    headers4 = ["S.No", "File Name", "Function Name"]
    widths4 = [6, 35, 45]
    rows4 = [(r['sno'], r['file'], r['function']) for r in func_rows]
    build_sheet(ws4, "Functions", headers4, widths4, rows4)

    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws0 = wb.create_sheet("Summary", 0)
    ws0.sheet_properties.tabColor = "1F3864"
    ws0.column_dimensions['A'].width = 28
    ws0.column_dimensions['B'].width = 18

    summary_data = [
        ("Total Files", len(file_records)),
        ("Python Files", sum(1 for r in file_records if r['type'] == 'Python')),
        ("Non-Python Files", sum(1 for r in file_records if r['type'] != 'Python')),
        ("", ""),
        ("Total Classes", len(class_rows)),
        ("Total Methods", len(method_rows)),
        ("Total Functions", len(func_rows)),
    ]

    ws0.row_dimensions[1].height = 40
    ws0.merge_cells("A1:B1")
    t = ws0.cell(row=1, column=1, value="PROJECT ANALYSIS SUMMARY")
    t.font = Font(bold=True, color="FFFFFF", name='Arial', size=14)
    t.fill = PatternFill("solid", start_color="1F3864")
    t.alignment = Alignment(horizontal='center', vertical='center')

    ws0.row_dimensions[2].height = 8

    highlight_bgs = ["2E75B6", "375623", "636363", "", "7030A0", "7030A0", "BF8F00"]

    for i, (label, value) in enumerate(summary_data, 3):
        ws0.row_dimensions[i].height = 28
        if label == "":
            continue
        bg = highlight_bgs[i - 3] if highlight_bgs[i - 3] else "F5F5F5"
        lc = ws0.cell(row=i, column=1, value=f"  {label}")
        lc.font = Font(bold=True, color="FFFFFF", name='Arial', size=10)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(vertical='center')
        lc.border = Border(bottom=Side(style='thin', color='FFFFFF'))

        vc = ws0.cell(row=i, column=2, value=value)
        vc.font = Font(bold=True, color="1F3864", name='Arial', size=14)
        vc.fill = PatternFill("solid", start_color="EBF0FA")
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.border = Border(bottom=Side(style='thin', color='D0D0D0'))

    ws0.freeze_panes = None

    wb.save(out_file)
    print(f"[OK] PROJECT_ANALYSIS.xlsx written ({len(file_records)} files, "
          f"{len(class_rows)} classes, {len(method_rows)} methods, {len(func_rows)} functions)")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    root = os.getcwd()
    print(f"\n[SCAN] Scanning: {root}\n")

    write_tree_txt(root, "PROJECT_TREE.txt")

    file_records = collect_files(root)
    print(f"[INFO] Found {len(file_records)} files")

    class_rows, method_rows, func_rows = collect_python_data(file_records, root)
    print(f"[INFO] Python: {len(class_rows)} classes | {len(method_rows)} methods | {len(func_rows)} functions")

    build_excel(file_records, class_rows, method_rows, func_rows, "PROJECT_ANALYSIS.xlsx")
    print("\n[SUCCESS] Files created:\n  - PROJECT_TREE.txt\n  - PROJECT_ANALYSIS.xlsx\n")


if __name__ == "__main__":
    main()