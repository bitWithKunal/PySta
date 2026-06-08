"""
Excel report writer for PySTA.
Generates Excel format timing reports.
"""

import pandas as pd
from typing import Dict, List, Optional, Any
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from Src.utils.logger import get_logger
from Src.utils.time_utils import TimeUtils

logger = get_logger(__name__)


class ExcelReportWriter:
    """Writes timing reports in Excel format."""

    def __init__(self):
        self.writer = None
        self.file_path = None

        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.violation_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.met_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.marginal_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_report(self, results: Dict[str, Any], file_path: str) -> bool:
        """
        Write comprehensive Excel report.

        Args:
            results: Analysis results dictionary
            file_path: Output file path

        Returns:
            True if successful, False otherwise
        """
        try:
            self.file_path = file_path

            # Create Excel writer
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                self.writer = writer

                # Write summary sheet
                self._write_summary_sheet(results)

                # Write setup paths sheet
                if 'setup_results' in results:
                    self._write_paths_sheet(results['setup_results'], 'Setup Paths', 'setup')

                # Write hold paths sheet
                if 'hold_results' in results:
                    self._write_paths_sheet(results['hold_results'], 'Hold Paths', 'hold')

                # Write violations sheet
                self._write_violations_sheet(results)

                # Write clock summary sheet
                if 'clock_constraints' in results:
                    self._write_clock_sheet(results['clock_constraints'])

                # Write graph statistics sheet
                if 'graph_builder' in results:
                    self._write_graph_sheet(results['graph_builder'])

                # Write OCV summary if available
                if 'ocv_analyzer' in results:
                    self._write_ocv_sheet(results['ocv_analyzer'])

                # Write SI summary if available
                if 'si_analyzer' in results:
                    self._write_si_sheet(results['si_analyzer'])

            # Post-process for formatting
            self._apply_formatting(file_path)

            logger.info(f"Excel report written to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to write Excel report: {e}", exc_info=True)
            return False

    def _write_summary_sheet(self, results: Dict[str, Any]):
        """Write summary sheet."""
        # Prepare summary data
        summary_data = []

        # General information
        summary_data.append(['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append([])

        # Design statistics
        if 'graph_builder' in results:
            graph_builder = results['graph_builder']
            summary_data.append(['Design Statistics', ''])
            summary_data.append(['Total Nodes', len(graph_builder.get_all_nodes())])
            summary_data.append(['Total Edges', graph_builder.edge_manager.get_num_edges()])
            summary_data.append(['Start Points', len(graph_builder.get_startpoints())])
            summary_data.append(['End Points', len(graph_builder.get_endpoints())])
            summary_data.append([])

        # Setup timing summary
        if 'setup_results' in results:
            setup = results['setup_results']
            summary_data.append(['Setup Timing', ''])
            summary_data.append(['Total Paths', setup.get('total_paths', 0)])
            summary_data.append(['Violations', setup.get('violations', 0)])
            summary_data.append(['Worst Slack (ps)', f"{setup.get('worst_slack', 0) * 1e12:.3f}"])
            summary_data.append(['TNS (ps)', f"{setup.get('tns', 0) * 1e12:.3f}"])
            summary_data.append(['WNS (ps)', f"{setup.get('wns', 0) * 1e12:.3f}"])
            summary_data.append([])

        # Hold timing summary
        if 'hold_results' in results:
            hold = results['hold_results']
            summary_data.append(['Hold Timing', ''])
            summary_data.append(['Total Paths', hold.get('total_paths', 0)])
            summary_data.append(['Violations', hold.get('violations', 0)])
            summary_data.append(['Worst Slack (ps)', f"{hold.get('worst_slack', 0) * 1e12:.3f}"])
            summary_data.append(['TNS (ps)', f"{hold.get('tns', 0) * 1e12:.3f}"])
            summary_data.append(['WNS (ps)', f"{hold.get('wns', 0) * 1e12:.3f}"])

        # Create DataFrame and write
        df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df.to_excel(self.writer, sheet_name='Summary', index=False)

    def _write_paths_sheet(self, results: Dict[str, Any], sheet_name: str,
                           analysis_type: str):
        """
        Write timing paths sheet.

        Args:
            results: Analysis results
            sheet_name: Excel sheet name
            analysis_type: 'setup' or 'hold'
        """
        paths = results.get('paths', [])

        if not paths:
            # Write empty sheet with message
            df = pd.DataFrame([['No paths found']], columns=['Message'])
            df.to_excel(self.writer, sheet_name=sheet_name, index=False)
            return

        # Prepare path data
        path_data = []
        for i, path in enumerate(paths[:1000]):  # Limit to 1000 paths
            path_data.append({
                'Path #': i + 1,
                'From': path.get('from', ''),
                'To': path.get('to', ''),
                'Clock': path.get('clock', ''),
                'Path Delay (ps)': path.get('delay', 0) * 1e12,
                'Required (ps)': path.get('required', 0) * 1e12,
                'Slack (ps)': path.get('slack', 0) * 1e12,
                'Stage Count': len(path.get('stages', [])),
                'Violation': 'Yes' if path.get('slack', 0) < 0 else 'No'
            })

        df = pd.DataFrame(path_data)
        df.to_excel(self.writer, sheet_name=sheet_name, index=False)

    def _write_violations_sheet(self, results: Dict[str, Any]):
        """Write violations sheet."""
        violations = []

        # Collect setup violations
        if 'setup_results' in results:
            for path in results['setup_results'].get('paths', []):
                if path.get('slack', 0) < 0:
                    path['type'] = 'setup'
                    violations.append(path)

        # Collect hold violations
        if 'hold_results' in results:
            for path in results['hold_results'].get('paths', []):
                if path.get('slack', 0) < 0:
                    path['type'] = 'hold'
                    violations.append(path)

        if not violations:
            df = pd.DataFrame([['No violations found']], columns=['Message'])
            df.to_excel(self.writer, sheet_name='Violations', index=False)
            return

        # Prepare violation data
        violation_data = []
        for i, path in enumerate(violations):
            violation_data.append({
                '#': i + 1,
                'Type': path.get('type', '').upper(),
                'From': path.get('from', ''),
                'To': path.get('to', ''),
                'Clock': path.get('clock', ''),
                'Slack (ps)': path.get('slack', 0) * 1e12,
                'Required (ps)': path.get('required', 0) * 1e12,
                'Arrival (ps)': path.get('arrival', 0) * 1e12,
                'Path Delay (ps)': path.get('delay', 0) * 1e12
            })

        # Sort by slack (most negative first)
        violation_data.sort(key=lambda x: x['Slack (ps)'])

        df = pd.DataFrame(violation_data)
        df.to_excel(self.writer, sheet_name='Violations', index=False)

    def _write_clock_sheet(self, clock_constraints):
        """Write clock summary sheet."""
        clocks = clock_constraints.get_all_clocks()

        if not clocks:
            df = pd.DataFrame([['No clocks defined']], columns=['Message'])
            df.to_excel(self.writer, sheet_name='Clocks', index=False)
            return

        clock_data = []
        for clock in clocks:
            clock_data.append({
                'Clock Name': clock.name,
                'Period (ps)': clock.get_period() * 1e12,
                'Frequency (MHz)': 1e6 / (clock.get_period() * 1e12),
                'Waveform Rise (ps)': clock.waveform[0] * 1e12,
                'Waveform Fall (ps)': clock.waveform[1] * 1e12,
                'Latency (ps)': clock.latency * 1e12,
                'Uncertainty (ps)': clock.uncertainty * 1e12,
                'Sources': ', '.join(clock.sources) if clock.sources else '',
                'Generated': 'Yes' if clock.is_generated else 'No'
            })

        df = pd.DataFrame(clock_data)
        df.to_excel(self.writer, sheet_name='Clocks', index=False)

    def _write_graph_sheet(self, graph_builder):
        """Write graph statistics sheet."""
        stats = [
            ['Metric', 'Value'],
            ['Total Nodes', len(graph_builder.get_all_nodes())],
            ['Total Edges', graph_builder.edge_manager.get_num_edges()],
            ['Start Points', len(graph_builder.get_startpoints())],
            ['End Points', len(graph_builder.get_endpoints())],
            ['Primary Inputs', sum(1 for n in graph_builder.get_all_nodes()
                                   if n.node_type.value == 'primary_input')],
            ['Primary Outputs', sum(1 for n in graph_builder.get_all_nodes()
                                    if n.node_type.value == 'primary_output')],
            ['Cell Inputs', sum(1 for n in graph_builder.get_all_nodes()
                                if n.node_type.value == 'cell_input')],
            ['Cell Outputs', sum(1 for n in graph_builder.get_all_nodes()
                                 if n.node_type.value == 'cell_output')]
        ]

        df = pd.DataFrame(stats[1:], columns=stats[0])
        df.to_excel(self.writer, sheet_name='Graph Statistics', index=False)

    def _write_ocv_sheet(self, ocv_analyzer):
        """Write OCV summary sheet."""
        summary = ocv_analyzer.get_ocv_summary()

        if not summary:
            df = pd.DataFrame([['No OCV data available']], columns=['Message'])
            df.to_excel(self.writer, sheet_name='OCV Analysis', index=False)
            return

        # Main statistics
        stats = [
            ['Metric', 'Value'],
            ['Analysis Mode', ocv_analyzer.analysis_mode],
            ['Paths Analyzed', summary.get('num_paths', 0)],
            ['Average Nominal Delay (ps)', summary.get('avg_nominal_delay', 0) * 1e12],
            ['Average Derated Delay (ps)', summary.get('avg_derated_delay', 0) * 1e12],
            ['Max Nominal Delay (ps)', summary.get('max_nominal_delay', 0) * 1e12],
            ['Max Derated Delay (ps)', summary.get('max_derated_delay', 0) * 1e12],
            ['Total OCV Penalty (ps)', summary.get('total_ocv_penalty', 0) * 1e12],
            ['Average Derate Ratio', summary.get('avg_derate_ratio', 1.0)]
        ]

        df_stats = pd.DataFrame(stats[1:], columns=stats[0])
        df_stats.to_excel(self.writer, sheet_name='OCV Analysis', index=False, startrow=0)

        # Derate factors
        derates = summary.get('derate_factors', {})
        if derates:
            derate_data = [[k, v] for k, v in derates.items()]
            df_derates = pd.DataFrame(derate_data, columns=['Factor', 'Value'])
            df_derates.to_excel(self.writer, sheet_name='OCV Analysis',
                                index=False, startrow=len(stats) + 2)

    def _write_si_sheet(self, si_analyzer):
        """Write SI analysis sheet."""
        summary = si_analyzer.get_si_summary() if hasattr(si_analyzer, 'get_si_summary') else {}

        if not summary:
            df = pd.DataFrame([['No SI data available']], columns=['Message'])
            df.to_excel(self.writer, sheet_name='SI Analysis', index=False)
            return

        stats = [
            ['Metric', 'Value'],
            ['Coupling Threshold', f"{summary.get('coupling_threshold', 0) * 100:.1f}%"],
            ['Noise Margin', f"{summary.get('noise_margin', 0) * 100:.1f}%"],
            ['Aggressors Analyzed', summary.get('aggressor_count', 0)],
            ['Average SI Penalty (ps)', summary.get('avg_penalty', 0) * 1e12],
            ['Max SI Penalty (ps)', summary.get('max_penalty', 0) * 1e12],
            ['Nets Affected', summary.get('nets_affected', 0)]
        ]

        df = pd.DataFrame(stats[1:], columns=stats[0])
        df.to_excel(self.writer, sheet_name='SI Analysis', index=False)

    def _apply_formatting(self, file_path: str):
        """
        Apply formatting to Excel file.

        Args:
            file_path: Path to Excel file
        """
        try:
            wb = openpyxl.load_workbook(file_path)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Apply header formatting
                for cell in ws[1]:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply conditional formatting based on sheet
                if sheet_name in ['Setup Paths', 'Hold Paths', 'Violations']:
                    self._format_path_sheet(ws)

                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

        except Exception as e:
            logger.error(f"Failed to apply Excel formatting: {e}")

    def _format_path_sheet(self, ws):
        """Apply formatting to path sheets."""
        # Find slack column
        slack_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Slack (ps)':
                slack_col = col
                break

        if slack_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=slack_col)
                try:
                    slack = float(cell.value)
                    if slack < 0:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = self.violation_fill
                    elif slack < 100:  # 100ps threshold
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = self.marginal_fill
                    else:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = self.met_fill
                except:
                    pass